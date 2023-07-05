from selenium import webdriver
from openpyxl import Workbook, load_workbook
from selenium.webdriver.common.by import By
import chromedriver_autoinstaller
import datetime
import time

date = datetime.date.today().strftime("%d-%m-%Y")
print(date)

chromedriver_autoinstaller.install()
driver = webdriver.Chrome()

ld_wb = load_workbook(r"E:\Durai\Scraping\Home_appliances\Web Url\Urls.xlsx")
ld_ws = ld_wb.active
print("No Of Rows : ", ld_ws.max_row)

wb = Workbook()
ws = wb.active

for r in range(2, ld_ws.max_row+1):

    if ld_ws.cell(row=r, column=6).value != 'N/A':
        # ws.cell(row=r, column=6).value = ld_ws.cell(row=r, column=6).value
        print("#" * 150)
        print("Darling retail")
        print('Range: ', r)
        print("Link : ", ld_ws.cell(row=r, column=6).value)

        driver.get(url= ld_ws.cell(row=r, column=6).value)
        time.sleep(2)

        

