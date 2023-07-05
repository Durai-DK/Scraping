import time, datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook, Workbook
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from Form.Product_Details import path
from selenium.common.exceptions import NoSuchElementException


s = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=s)


date = datetime.date.today().strftime("%d-%m-%Y")
print(date)


new_wb = Workbook()
new_ws = new_wb.active


class PriceCompression:
    def __init__(self,  row_num=None, ws=None):
        self.row_num = row_num
        self.ws = ws

    def heading(self):

        new_ws.cell['a1'] = "Item Code"
        new_ws.cell['a2'] = "Model Name"
        new_ws.cell['a3'] = "Poorvika Price"
        new_ws.cell['a4'] = "Flipkart Url"
        new_ws.cell['a5'] = "Flipkart Price"
        new_ws.cell['a6'] = "Amazon Url"
        new_ws.cell['a7'] = "Amazon Price"
        new_ws.cell['a8'] = "Croma Url"
        new_ws.cell['a9'] = "Croma Price"
        new_ws.cell['a10'] = "Vijay Sale Url"
        new_ws.cell['a11'] = "Vijay Sale Price"
        new_ws.cell['a12'] = "Reliance Digital Name"
        new_ws.cell['a13'] = "Reliance Digital Price"

# -----------------------------------------------------------------------------------------------------------------
    def flipkart(self):

        if self.ws.cell(row=self.row_num, column=4).value != "N/A":
            print("#" * 100)
            print("Web Site : Flipkart ")
            print("Web Link :", self.ws.cell(row=self.row_num, column=4).value)
            new_ws.cell(row=self.row_num, column=4).value = self.ws.cell(row=self.row_num, column=4).value

            try:
                driver.get(url=self.ws.cell(row=self.row_num, column=4).value)
                time.sleep(1)

                price = driver.find_element(By.CLASS_NAME, "_30jeq3").text[1:]

                new_ws.cell(row=self.row_num, column=5).value = price
                print("Product Price :", price)

            except:
                pass

# -----------------------------------------------------------------------------------------------------------------
    def amazon(self):
        if self.ws.cell(row=self.row_num, column=5).value != "N/A":
            print("#" * 100)
            print("Web Site :Amazon")
            print("Web Link :", self.ws.cell(row=self.row_num, column=5).value)
            new_ws.cell(row=self.row_num, column=6).value = self.ws.cell(row=self.row_num, column=5).value
            try:
                driver.get(url=self.ws.cell(row=self.row_num, column=5).value)
                driver.implicitly_wait(3)

                try:
                    price = driver.find_element(By.ID, "apex_desktop")
                    price1 = price.find_element(By.CLASS_NAME, "a-price-whole")
                    print("Amazon Price 1 :", price1.text)
                    new_ws.cell(row=self.row_num, column=7).value = price1.text

                except NoSuchElementException:
                    pric = driver.find_element(By.ID, "apex_desktop")
                    pric2 = pric.find_element(By.CLASS_NAME, "apexPriceToPay")
                    print("Amazon Price 2 :", pric2.text)
                    new_ws.cell(row=self.row_num, column=7).value = pric2.text

            except:
                pass

# ---------------------------------------------------------------------------------------------------------------------
    def croma(self):
        # time.sleep(2)
        if self.ws.cell(row=self.row_num, column=6).value != "N/A":
            print("#" * 100)
            print("Web Site :Croma")
            print("Web Link :", self.ws.cell(row=self.row_num, column=6).value)
            new_ws.cell(row=self.row_num, column=8).value = self.ws.cell(row=self.row_num, column=6).value

            try:
                driver.get(url=self.ws.cell(row=self.row_num, column=6).value)
                time.sleep(3)

                for price in driver.find_elements(By.CLASS_NAME, "outer-product-pricebox"):
                    price1 = price.find_element(By.ID, "pdp-product-price").text
                    print("Croma Price 1 = ", price1[1:])
                    new_ws.cell(row=self.row_num, column=9).value = price1[1:]

            except:
                pass

# ---------------------------------------------------------------------------------------------------------------------
    def vijay_sale(self):
        if self.ws.cell(row=self.row_num, column=7).value != "N/A":
            print("#" * 100)
            print("Web Site :vijay sale")
            print("Web Link :", self.ws.cell(row=self.row_num, column=7).value)
            new_ws.cell(row=self.row_num, column=10).value = self.ws.cell(row=self.row_num, column=7).value

            try:
                driver.get(url=self.ws.cell(row=self.row_num, column=7).value)
                time.sleep(2)

                try:
                    print("Vijay Sale Price 1 :", driver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_fillprice"]/div[1]/div[1]/span[2]/span').text)
                    price = driver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_fillprice"]/div[1]/div[1]/span[2]/span').text
                    new_ws.cell(row=self.row_num, column=11).value = price

                except NoSuchElementException:
                    print("Vijay Sale Price 2 :", driver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_fillprice"]/div/span[2]/span').text)
                    price = driver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_fillprice"]/div/span[2]/span').text
                    new_ws.cell(row=self.row_num, column=11).value = price

            except:
                pass
# ---------------------------------------------------------------------------------------------------------------------
    def reliance(self):
        if self.ws.cell(row=self.row_num, column=8).value != "N/A":
            print("#" * 100)
            print("Web Site :Reliance")
            print("Web Link :", self.ws.cell(row=self.row_num, column=8).value)
            new_ws.cell(row=self.row_num, column=12).value = self.ws.cell(row=self.row_num, column=8).value

            try:
                driver.get(url=self.ws.cell(row=self.row_num, column=8).value)
                time.sleep(1)

                for pric in driver.find_elements(By.CLASS_NAME, 'pdp__priceSection__priceListText'):
                    price = pric.find_element(By.CLASS_NAME, "kFBgPo").text
                    print("Reliance Price 1 = ", price[1:])
                    new_ws.cell(row=self.row_num, column=13).value = price[1:]

            except:
                pass

#######################################################################################################################
class RunCompression:

    def run__all(self, **kwargs):
        if kwargs.get('head') == "Accessories":
            wb = load_workbook(path["Accessories"]['Url'] + date + ".xlsx")
            ws = wb.active

        elif kwargs.get('head') == "Laptop":
            wb = load_workbook(path["Laptop"]['Url'] + date + ".xlsx")
            ws = wb.active

        elif kwargs.get('head') == "Mobile":
            wb = load_workbook(path["Mobile"]['Url'] + date + ".xlsx")
            ws = wb.active

        elif kwargs.get('head') == "Tv":
            wb = load_workbook(path["Tv"]['Url'] + date + ".xlsx")
            ws = wb.active

        elif kwargs.get('head') == "Tablets":
            wb = load_workbook(path["Tablets"]['Url'] + date + ".xlsx")
            ws = wb.active

        elif kwargs.get('head') == "KA":
            wb = load_workbook(path["KA"]['Url'] + date + ".xlsx")
            ws = wb.active

        ph = PriceCompression()
        ph.heading()

        for r in range(kwargs.get('start'), kwargs.get('end')+1):
            print("")
            print(r)
            new_ws.cell(row=r, column=1).value = ws.cell(row=r, column=2).value
            new_ws.cell(row=r, column=2).value = ws.cell(row=r, column=1).value
            new_ws.cell(row=r, column=3).value = ws.cell(row=r, column=3).value
            pc = PriceCompression(row_num=r, ws=ws)
            pc.flipkart()
            pc.amazon()
            pc.croma()
            pc.vijay_sale()
            pc.reliance()
            self.save(head=kwargs.get('head'), path=kwargs.get('path'))
        driver.quit()

#######################################################################################################################

    def save(self, **kwargs):

        if kwargs.get("head") == "Accessories":
            path_save = path["Accessories"]['Save']
            path_num = str(kwargs.get('path'))
            new_wb.save(path_save + path_num + " Scraping All " + date + ".xlsx")

        elif kwargs.get("head") == "Mobile":
            path_save = path["Mobile"]['Save']
            path_num = str(kwargs.get('path'))
            new_wb.save(path_save + path_num + " Scraping All " + date + ".xlsx")

        elif kwargs.get("head") == "KA":
            path_save = path["KA"]['Save']
            path_num = str(kwargs.get('path'))
            new_wb.save(path_save + path_num + " Scraping All " + date + ".xlsx")

        elif kwargs.get("head") == "Laptop":
            new_wb.save(path["Laptop"]['Save'] + date + ".xlsx")

        elif kwargs.get("head") == "Tv":
            new_wb.save(path["Tv"]['Save'] + date + ".xlsx")

        elif kwargs.get("head") == "Tablets":
            new_wb.save(path["Tablets"]['Save'] + date + ".xlsx")

#######################################################################################################################
