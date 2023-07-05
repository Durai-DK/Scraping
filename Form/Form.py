from selenium.webdriver.common.by import By
import datetime
from selenium import webdriver
from openpyxl import Workbook
import time
import chromedriver_autoinstaller
# from selenium.webdriver.chrome.service import Service as ChromeService
# from webdriver_manager.chrome import ChromeDriverManager

date = datetime.datetime.now().strftime("%d-%m-%Y")
# s = ChromeService(ChromeDriverManager().install())
# driver = webdriver.Chrome(service=s)
# chromedriver_autoinstaller.install()
driver = webdriver.Chrome()


wb = Workbook()
ws = wb.active

########################################################################################################################

class Model:

    def __init__(self, url, row_num, heading):

        self.url = url
        self.row_num = row_num
        self.heading = heading

########################################################################################################################

    def product_page(self):

        for box in driver.find_elements(By.CLASS_NAME, "product-cardlist_card__description__4m5gI"):
            name = box.find_element(By.TAG_NAME, "b").text
            price = box.find_element(By.CLASS_NAME, "whitespace-nowrap").text
            link = box.find_element(By.TAG_NAME, "a").get_attribute("href")

            print("Name : ", name)
            print("Price : ", price[2:])
            print("Urls : ", link)
            print(" ")

            ws.cell(row=self.row_num, column=1).value = name
            ws.cell(row=self.row_num, column=2).value = price[2:]
            ws.cell(row=self.row_num, column=3).value = link
            ws.cell(row=self.row_num, column=4).value = self.heading

            self.product_save(head=self.heading)
            self.row_num = self.row_num + 1

########################################################################################################################
    def product_page_list(self, **kwargs):
        print("Title :", kwargs.get('head'))
        page = ""

        if kwargs.get('head') == "Mobile & Accessories":
            page = 23
            # page = 2

        elif kwargs.get('head') == "Computer & Laptop Accessories":
            page = 6
            # page = 2

        elif kwargs.get('head') == "Tab & Ipad Accessories":
            page = 5
            # page = 2

        elif kwargs.get('head') == "Audio Accessories":
            page = 10
            # page = 2

        elif kwargs.get('head') == "Smart Technology":
            page = 10
            # page = 2

        elif kwargs.get('head') == "Laptops":
            page = 10
            # page = 2

        elif kwargs.get('head') == "Tablets":
            page = 10
            # page = 2

        elif kwargs.get('head') == "Mobiles":
            page = 30
            # page = 2

        elif kwargs.get('head') == "Tv":
            page = 5
            # page = 2

        elif kwargs.get('head') == "Kitchen Appliances":
            page = 20
            # page = 2

        self.product(page=page)

########################################################################################################################

    def product(self, **kwargs):
        print(kwargs.get('page'))
        # try:
        for r in range(1, int(kwargs.get('page'))):
            print("                     ")
            print("Page range:" + str(r))
            driver.get(self.url + str(r))
            time.sleep(5)
            self.product_page()
        # except:
        #     pass

        print(self.heading + " Complete")
        print("#" * 60)

########################################################################################################################
    def product_1(self):
        for r in range(1,2):
            print(r)

            driver.get(self.url)

            for box in driver.find_elements(By.CLASS_NAME, "product-cardlist_card__description__4m5gI"):
                name = box.find_element(By.TAG_NAME, "b").text
                price = box.find_element(By.CLASS_NAME, "whitespace-nowrap").text
                link = box.find_element(By.TAG_NAME, "a").get_attribute("href")

                print(name)
                print(price[2:])
                print(link)
                print(" ")

                ws.cell(row=self.row_num, column=1).value = name
                ws.cell(row=self.row_num, column=2).value = price[2:]
                ws.cell(row=self.row_num, column=3).value = link
                ws.cell(row=self.row_num, column=4).value = self.heading

                self.row_num = self.row_num + 1
                self.product_save(head=self.heading)
            print(self.heading + " Complete")
            print("#" * 60)

########################################################################################################################

    def product_num(self):
        return self.row_num

########################################################################################################################

    def product_save(self,**kwargs):

        if kwargs.get('head') == "Laptops":
            wb.save(r"E:\Durai\Scraping\Laptop\Save Data's\Poorvika Files\Laptop " + date + ".xlsx")

        elif kwargs.get('head') == "Tablets":
            wb.save(r"E:\Durai\Scraping\Tablets\Save Data's\Poorvika Files\Tablets " + date + ".xlsx")

        elif kwargs.get('head') == "Tv":
            wb.save(r"E:\Durai\Scraping\Tv\Save Data's\Poorvika Files\Tv " + date + ".xlsx")

        elif kwargs.get('head') == "Mobiles":
            wb.save(r"E:\Durai\Scraping\Mobile\Save Data's\Poorvika Files\Mobiles " + date + ".xlsx")

        elif kwargs.get('head') == "Kitchen Appliances":
            wb.save(r"E:\Durai\Scraping\Kitchen Appliances\Save Data's\Poorvika Files\kitchen Appliances " + date + ".xlsx")

        else:
            wb.save(r"E:\Durai\Scraping\Accessories\Save Data's\Poorvika Files\Accessories " + date + ".xlsx")

########################################################################################################################
