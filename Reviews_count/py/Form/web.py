import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook,Workbook
import datetime
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService

date = datetime.datetime.now().strftime("%d-%m-%Y")
s = ChromeService(ChromeDriverManager().install())
driver = webdriver.Chrome(service=s)

new_wb = Workbook()
new_ws = new_wb.active

wb = load_workbook(r"E:\Durai\Scraping/Reviews_count/Urls/Google Ratings & Reviews Urls.xlsx")
ws = wb.active

new_ws.cell(row=1,column=1).value = "APX"
new_ws.cell(row=1,column=2).value = "Links"
new_ws.cell(row=1,column=3).value = "Ratings"
new_ws.cell(row=1,column=4).value = "Reviews"


class GoogleRating:
    def __init__(self, row_num):
        self.row_num = row_num

    def Scrap(self):

        try:
            driver.get(url=ws.cell(row=self.row_num, column=2).value)
            print(ws.cell(row=self.row_num, column=2).value)
            time.sleep(2)

            Ratings = driver.find_element(By.CLASS_NAME, "F7nice").text
            print("Ratings = ", Ratings[:4])
            new_ws.cell(row=self.row_num, column=3).value = Ratings[:4]

            Reviews = driver.find_element(By.XPATH, "//*[@id='QA0Szd']/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div/div[1]/div[2]/div/div[1]/div[2]/span[2]/span/span").text
            print("Reviews = ", Reviews[1:-1])
            new_ws.cell(row=self.row_num, column=4).value = Reviews[1:-1]

        except:
            pass

    def save(self,**kwargs):

        new_ws.cell(row=self.row_num, column=1).value = ws.cell(row=self.row_num,column=1).value
        new_ws.cell(row=self.row_num, column=2).value = ws.cell(row=self.row_num,column=2).value

        path = r"D:\Mohi\Scraping\Reviews_Count\Save Fiels\Scraping Fiels\GMB R&R "
        path_num = str(kwargs.get("path"))
        new_wb.save(path + path_num + " " + date + ".xlsx")



class GMB:

    def run(self,**kwargs):

        for r in range(kwargs.get("start"),kwargs.get("end")+1):
            print(" ")
            print(r)

            dk = GoogleRating(row_num=r)

            dk.Scrap()
            dk.save(path=kwargs.get("path"))

        driver.quit()