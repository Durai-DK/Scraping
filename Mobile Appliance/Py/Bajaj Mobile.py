from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from openpyxl import Workbook,load_workbook
import datetime
import time


date = datetime.datetime.now().strftime("%d-%m-%Y")
s = ChromeService(ChromeDriverManager().install())
driver = webdriver.Chrome(service=s)

driver.get("https://www.google.com/")

l_wb = load_workbook(r"D:\Durai\Scraping\Mobile Appliance\Urls\Mobile Appliance.xlsx")
l_ws = l_wb.active
print("No Of Rows",l_ws.max_row)

wb = Workbook()
ws = wb.active

ws["a1"] = "Item Code"
ws["b1"] = "Bajaj Url"
ws["c1"] = "Model"
ws["d1"] = "Poorvika Price"
ws["e1"] = "Bajaj Price"


for r in range(2,l_ws.max_row+1):

    ws.cell(row=r,column=1).value = l_ws.cell(row=r,column=1).value
    ws.cell(row=r,column=3).value = l_ws.cell(row=r,column=2).value

    if l_ws.cell(row=r,column=4).value != "N/A":
        ws.cell(row=r,column=2).value = l_ws.cell(row=r,column=4).value

        print("#" * 120)
        print(" ")
        print("Bajaj")
        print('Range : ', r)
        print("Bajaj Url : ", l_ws.cell(row=r, column=4).value)

        try:
            driver.get(url=l_ws.cell(row=r, column=4).value)
            time.sleep(2)

            for box in driver.find_elements(By.CLASS_NAME, "priceDetails"):
                price = box.find_element(By.TAG_NAME, 'h3').text
                print("Bajaj Price = ", price[2:-22])
                ws.cell(row=r,column=5).value = price[2:-22]

        except:
            pass

    wb.save(r"D:\Durai\Scraping\Mobile Appliance\Save Data\Scraping Files\Total Scraping\Mobile Bajaj " + date + ".xlsx")

driver.quit()
