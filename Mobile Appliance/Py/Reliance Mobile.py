from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from openpyxl import Workbook,load_workbook
import datetime
import time


date = datetime.datetime.now().strftime("%d-%m-%Y")
s = ChromeService(ChromeDriverManager().install())
driver = webdriver.Chrome(service=s)

driver.get("https://www.google.com/")

l_wb = load_workbook(r"D:\Durai\Scraping\Mobile Appliance\Urls\Mobile Appliance.xlsx")
l_ws = l_wb.active
print("No Of Rows",l_ws.max_row)

wb = Workbook()
ws = wb.active

ws["a1"] = "Item Code"
ws["b1"] = "Reliance Url"
ws["c1"] = "Model"
ws["d1"] = "Poorvika Price"
ws["e1"] = "Reliance Price"

for r in range(2,l_ws.max_row+1):

    ws.cell(row=r,column=1).value = l_ws.cell(row=r,column=1).value
    ws.cell(row=r,column=3).value = l_ws.cell(row=r,column=2).value

    if l_ws.cell(row=r,column=8).value != "N/A":
        ws.cell(row=r,column=2).value = l_ws.cell(row=r,column=8).value

        print("#" * 120)
        print(" ")
        print("Reliance")
        print('Range : ', r)
        print("Reliance Url : ", l_ws.cell(row=r, column=8).value)

        try:
            driver.get(url=l_ws.cell(row=r, column=8).value)
            time.sleep(2)

            for pri in driver.find_elements(By.CLASS_NAME, 'pdp__priceSection__priceListText'):
                price = pri.find_element(By.CLASS_NAME, "kFBgPo").text
                print("Reliance Price 1 = ", price[1:])
                ws.cell(row=r, column=5).value = price[1:]

        except:
            pass

    wb.save(r"D:\Durai\Scraping\Mobile Appliance\Save Data\Scraping Files\Total Scraping\Mobile Reliance " + date + ".xlsx")

driver.quit()
