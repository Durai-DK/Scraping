from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from openpyxl import Workbook,load_workbook
import datetime
import time


date = datetime.datetime.now().strftime("%d-%m-%Y")
s = ChromeService(ChromeDriverManager().install())
driver = webdriver.Chrome(service=s)

driver.get("https://www.google.com/")

l_wb = load_workbook(r"D:\Durai\Scraping\Mobile Appliance\Urls\Mobile Appliance.xlsx")
l_ws = l_wb.active
print("No Of Rows",l_ws.max_row)

wb = Workbook()
ws = wb.active

ws["a1"] = "Item Code"
ws["b1"] = "Chennai Url"
ws["c1"] = "Model"
ws["d1"] = "Poorvika Price"
ws["e1"] = "Chennai Price"

for r in range(2,l_ws.max_row+1):

    ws.cell(row=r,column=1).value = l_ws.cell(row=r,column=1).value
    ws.cell(row=r, column=3).value = l_ws.cell(row=r, column=2).value

    if l_ws.cell(row=r,column=5).value != "N/A":
        ws.cell(row=r,column=2).value = l_ws.cell(row=r,column=5).value

        print("#" * 120)
        print(" ")
        print("Chennai")
        print('Range : ', r)
        print("Chennai Url : ", l_ws.cell(row=r, column=5).value)

        try:
            driver.get(url=l_ws.cell(row=r, column=5).value)
            time.sleep(2)

            for box in driver.find_elements(By.CLASS_NAME, "price-box"):
                price = box.find_element(By.CLASS_NAME, "special-price").text
                print("Chennai Price = ", price[4:])
                ws.cell(row=r, column=5).value = price[4:]

        except:
            pass

    wb.save(r"D:\Durai\Scraping\Mobile Appliance\Save Data\Scraping Files\Total Scraping\Mobile Chennai " + date + ".xlsx")

driver.quit()
