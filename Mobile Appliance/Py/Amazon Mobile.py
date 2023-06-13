from selenium.webdriver.common.by import By
from openpyxl import Workbook,load_workbook
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService
import datetime
import time

s = ChromeService(ChromeDriverManager().install())
driver = webdriver.Chrome(service=s)

date = datetime.datetime.now().strftime("%d-%m-%Y")

driver.get("https://www.google.com/")

l_wb = load_workbook(r"D:\Durai\Scraping\Mobile Appliance\Urls\Mobile Appliance.xlsx")
l_ws = l_wb.active
print("No Of Rows",l_ws.max_row)

wb = Workbook()
ws = wb.active

ws["a1"] = "Item Code"
ws["b1"] = "Amazon Url"
ws["c1"] = "Model"
ws["d1"] = "Poorvika Price"
ws["e1"] = "Amazon Price"


for r in range(2,l_ws.max_row+1):

    ws.cell(row=r,column=1).value = l_ws.cell(row=r,column=1).value
    ws.cell(row=r,column=3).value = l_ws.cell(row=r,column=2).value


    if l_ws.cell(row=r,column=3).value != "N/A":
        ws.cell(row=r,column=2).value = l_ws.cell(row=r,column=3).value

        print("#" * 120)
        print(" ")
        print("Amazon")
        print('Range : ', r)
        print("Amazon Url : ", l_ws.cell(row=r, column=3).value)

        try:
            driver.get(url=l_ws.cell(row=r, column=3).value)
            time.sleep(2)

            try:
                for price1 in driver.find_elements(By.ID, "olp_feature_div"):
                    for price2 in price1.find_elements(By.CLASS_NAME, "a-size-base"):
                        print("Amazon Price 3 = ", price2.text[1:])
                        ws.cell(row=r, column=5).value = price2.text[1:]
            except:
                pass

            try:
                price = driver.find_element(By.ID, "apex_desktop")
                price1 = price.find_element(By.CLASS_NAME, "a-price-whole")
                print("Amazon Price 2 = ", price1.text)
                ws.cell(row=r, column=5).value = price1.text
            except:
                pass

            try:
                price = driver.find_element(By.ID, "apex_desktop")
                price2 = price.find_element(By.CLASS_NAME, "apexPriceToPay")
                print("Amazon Price 1 = ", price2.text[1:])
                ws.cell(row=r, column=5).value = price2.text[1:]
            except:
                pass

        except:
            pass

    wb.save(r"D:\Durai\Scraping\Mobile Appliance\Save Data\Scraping Files\Total Scraping\Mobile Amazon " + date + ".xlsx")

driver.quit()