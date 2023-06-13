from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from openpyxl import Workbook,load_workbook
import datetime
import time


date = datetime.datetime.now().strftime("%d-%m-%Y")
s = ChromeService(ChromeDriverManager().install())
driver = webdriver.Chrome(service=s)

driver.get("https://www.google.com/")

l_wb = load_workbook(r"D:\Durai\Scraping\Mobile Appliance\Urls\Mobile Appliance.xlsx")
l_ws = l_wb.active
print("No Of Rows",l_ws.max_row)

wb = Workbook()
ws = wb.active

ws["a1"] = "Item Code"
ws["b1"] = "Croma Url"
ws["c1"] = "Model"
ws["d1"] = "Poorvika Price"
ws["e1"] = "Croma Price"

for r in range(2,l_ws.max_row+1):
# for r in range(1821,l_ws.max_row+1):

    ws.cell(row=r,column=1).value = l_ws.cell(row=r,column=1).value
    ws.cell(row=r,column=3).value = l_ws.cell(row=r,column=2).value


    if l_ws.cell(row=r,column=6).value != "N/A":
        ws.cell(row=r,column=2).value = l_ws.cell(row=r,column=6).value

        print("#" * 120)
        print(" ")
        print("Croma")
        print('Range : ', r)
        print("Croma Url : ", l_ws.cell(row=r, column=6).value)

        try:
            driver.get(url=l_ws.cell(row=r, column=6).value)
            time.sleep(3)

            for price in driver.find_elements(By.CLASS_NAME, "outer-product-pricebox"):
                price1 = price.find_element(By.ID, "pdp-product-price").text
                print("Croma Price 1 = ", price1[1:])
                ws.cell(row=r, column=5).value = price1[1:]

                if price1[1:] == " ":

                    for pric in driver.find_elements(By.CLASS_NAME, "outer-product-pricebox"):
                        for pric2 in pric.find_elements(By.CLASS_NAME, "new-price"):
                            pric3 = pric2.find_element(By.CLASS_NAME, "amount").text
                            print("Croma Price 2 = ", pric3[1:])
                            ws.cell(row=r, column=5).value = pric3[1:]

        except:
            pass

    wb.save(r"D:\Durai\Scraping\Mobile Appliance\Save Data\Scraping Files\Total Scraping\Mobile Croma " + date + ".xlsx")

driver.quit()
