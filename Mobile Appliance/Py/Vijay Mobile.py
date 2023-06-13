from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from openpyxl import Workbook,load_workbook
import datetime
import time


date = datetime.datetime.now().strftime("%d-%m-%Y")
s = ChromeService(ChromeDriverManager().install())
driver = webdriver.Chrome(service=s)


l_wb = load_workbook(r"D:\Durai\Scraping\Mobile Appliance\Urls\Mobile Appliance.xlsx")
l_ws = l_wb.active
print("No Of Rows",l_ws.max_row)

driver.get("https://www.google.com/")

wb = Workbook()
ws = wb.active

ws["a1"] = "Item Code"
ws["b1"] = "Vijay Url"
ws["c1"] = "Model"
ws["d1"] = "Poorvika Price"
ws["e1"] = "Vijay Price"

for r in range(2,l_ws.max_row+1):
# for r in range(697,l_ws.max_row+1):

    ws.cell(row=r,column=1).value = l_ws.cell(row=r,column=1).value
    ws.cell(row=r,column=3).value = l_ws.cell(row=r,column=2).value

    if l_ws.cell(row=r,column=11).value != "N/A":
        ws.cell(row=r,column=2).value = l_ws.cell(row=r,column=11).value

        print("#" * 120)
        print(" ")
        print("Vijay")
        print('Range : ', r)
        print("Vijay Url : ", l_ws.cell(row=r, column=11).value)

        driver.get(url=l_ws.cell(row=r, column=11).value)
        time.sleep(2)
        try:

            try:

                price1 = driver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_fillprice"]/div[1]/div[1]/span[2]/span').text
                print("Vijay Price 1 = ", price1)
                ws.cell(row=r, column=5).value = price1

            except:

                price2 = driver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_fillprice"]/div/span[2]/span').text
                print("Vijay Price 2 = ", price2)
                ws.cell(row=r, column=5).value = price2

        except:
            pass

    wb.save(r"D:\Durai\Scraping\Mobile Appliance\Save Data\Scraping Files\Total Scraping\Mobile Vijay " + date + ".xlsx")

driver.quit()
