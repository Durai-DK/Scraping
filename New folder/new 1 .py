from selenium.webdriver.common.by import By
import datetime
from selenium import webdriver
from openpyxl import Workbook,load_workbook
import time
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service

date = datetime.date.today().strftime("%d-%m-%Y")
print(date)

s = Service(ChromeDriverManager().install())

driver = webdriver.Chrome(service=s)

l_wb = load_workbook(r"E:\Durai\Scraping\New folder\Kitchen Appliances 03-07-2023.xlsx")
l_ws = l_wb.active

wb = Workbook()
ws = wb.active

for r in range(330, l_ws.max_row+1):
# for r in range(1, 5):
    print("")
    print(r)

    driver.get(url=l_ws.cell(row=r, column=2).value)
    time.sleep(3)

    name = driver.find_element(By.CLASS_NAME, "center-content_product_name__GvSMf").text
    item_Code = driver.find_element(By.CLASS_NAME, "style_item_code__nBTcg").text

    print(name)
    print(item_Code[15:-1])

    ws.cell(row=r, column=1).value = l_ws.cell(row=r, column=1).value
    ws.cell(row=r, column=2).value = item_Code[15:-1]

wb.save((r"E:\Durai\Scraping\New folder\Kitchen Appliances item code 2 " + date +".xlsx"))
